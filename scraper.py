from selenium import webdriver
import time
from datetime import datetime
from selenium.webdriver.common.by import By
import openpyxl

# calculate the time necessary to execute the script
start = time.time()

# add the timestamp to the xls
now = datetime.now()
now_str = now.strftime("%d/%m/%Y %H:%M:%S")

# load xls
wb = openpyxl.load_workbook('prod.xlsx')
ws = wb.active

# locate the first empty row
first_empty_row = ws.max_row + 1
ws.cell(row=first_empty_row, column=1).value = now_str

# category products from site
categories = ['https://myebox.ro/categorii/accesorii-de-impachetat/',
              'https://myebox.ro/categorii/benzi-adezive/',
              'https://myebox.ro/categorii/cutii-protectie-sticle/',
              'https://myebox.ro/categorii/cutii-de-carton/',
              'https://myebox.ro/categorii/fixare-si-umplere/fulgi-umplutura-biodegradabili/',
              'https://myebox.ro/categorii/fixare-si-umplere/hartie-tip-fagure/',
              'https://myebox.ro/categorii/fixare-si-umplere/hartie-cadouri/',
              'https://myebox.ro/categorii/fixare-si-umplere/hartie-umplutura/',
              'https://myebox.ro/categorii/folie-cu-bule/folie-cu-bule-clasica/',
              'https://myebox.ro/categorii/folie-cu-bule/folie-cu-bule-ecologica/',
              'https://myebox.ro/categorii/folie-stretch/',
              'https://myebox.ro/categorii/folii-autoadezive/',
              'https://myebox.ro/categorii/pungi-ziplock/',
              'https://myebox.ro/categorii/plicuri/plicuri-awb-autoadezive/'
              'https://myebox.ro/categorii/plicuri/plicuri-cu-bule/',
              'https://myebox.ro/categorii/plicuri/plicuri-corespondenta/',
              'https://myebox.ro/categorii/plicuri/plicuri-curierat/',
              'https://myebox.ro/categorii/plicuri/plicuri-curierat-eco/',
              'https://myebox.ro/categorii/plicuri/plicuri-carton/']


browser = webdriver.Chrome()

for category in categories:
    browser.get(category)
    time.sleep(3)
    height = browser.execute_script("return document.body.scrollHeight;")
    print(height)
    print("CATEGORY: {}".format(category))
    # constructs a set with all the links found on page
    links = set()
    for x in range(10):
        browser.execute_script("window.scrollTo(0,document.body.scrollHeight)")
        time.sleep(3)
        new_height = browser.execute_script("return document.body.scrollHeight;")
        print(new_height)
        if new_height == height:
            break
        height = new_height
        products = browser.find_elements(By.CSS_SELECTOR, "div > div > h3 > a")
        for i in products:
            links.add(i.get_attribute("href"))

    print(len(links))
    browser.refresh()
    time.sleep(2)

    # constructs a list with all the column headers found on the xls file
    header_list = []
    for header in ws[1]:
        header_list.append(header.value)

    # if link not found in header list, add a new column and the header
    for link in links:
        if link in header_list:
            continue
        else:
            ws.insert_cols(idx=ws.max_column + 1)
            ws.cell(row=1, column=ws.max_column+1).value = link

    # to be able to find the column index, construct a dictionary wity header as key and index as value
    column_names = {}
    current = 0
    for column in ws.iter_cols(1, ws.max_column):
        column_names[column[0].value] = current
        current += 1

    for x in links:
        browser.get(x)
        time.sleep(1)
        try:
            # time.sleep(3)
            print(x)
            b = browser.find_element(By.XPATH, "(//p[@class='stock in-stock'])[1]")
            try:
                stoc = browser.find_element(By.XPATH, "//input[@title='Cantitate']").get_attribute("max")
                if stoc != "":
                    print(stoc)
                    ws.cell(row=first_empty_row, column=column_names[x] + 1).value = stoc
                else:
                    c = browser.find_element(By.XPATH, "//p[@class='stock in-stock']")
                    stoc = c.get_attribute("innerHTML").split("Doar ")[1].split(" ")[0]
                    print(stoc)
                    ws.cell(row=first_empty_row, column=column_names[x] + 1).value = stoc
            except:
                # x = browser.find_element(By.XPATH, "//p[@class='stock out-of-stock']")
                # stoc = b.get_attribute("class").lstrip("stock").strip()
                stoc = "DATA NOT AVAILABLE"
                print(stoc)
                ws.cell(row=first_empty_row, column=column_names[x] + 1).value = "N\\A"
        except:
            print("N\\A")
            ws.cell(row=first_empty_row, column=column_names[x] + 1).value = "N\\A"

# save the xls
wb.save('prod.xlsx')

# calculate the time necessary to execute the script
end = time.time()
total_time = end - start
print("\n TOTAL TIME: {} minutes".format(int(total_time/60)))